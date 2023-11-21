import openpyxl.utils.exceptions as openpyxl_exceptions
import datetime
from openpyxl import load_workbook


def entrypoint():
    wb = read_file()
    while wb is None:
        print("read_file failed, try again...")
        wb = read_file()
    print_sheets(wb)
    sheet = select_sheet(wb)
    while sheet is None:
        sheet = select_sheet(wb)
    return parse_for_data(sheet)


def read_file():
    """
    Prints a prompt for user
    Receives string input
    Validates input
    Returns
    """
    print("Enter the name of the mileage sheet. Don't forget the extension: ")
    user_input_filename = input()
    try:
        workbook = load_workbook(filename=user_input_filename, data_only=True)
    except FileNotFoundError:
        print("File not found")
        return None
    except openpyxl_exceptions.InvalidFileException:
        print("Unsupported file format")
        return None
    else:
        print("File loaded successfully")
        return workbook


def print_sheets(workbook):
    print('Mileage sheets: ')
    print(workbook.sheetnames)


def select_sheet(workbook):
    """
    Selects page
    :param workbook:
    :return:
    """
    print("Which sheet would you like to use?")
    user_input = input()
    if user_input in workbook.sheetnames:
        print("Copying " + user_input)
        current_worksheet = workbook[user_input]
        return current_worksheet
    else:
        print("The sheet: " + user_input + " was not found.")
        return None


def parse_for_data(current_worksheet):
    """
    Parses a worksheet for data, returns a list of data
    """
    list_of_rows = []
    for x in range(8, 35):
        temp_list = []
        for row in current_worksheet[x]:
            if row.value is not None:
                if type(row.value) is datetime.datetime:
                    date_format = row.value.strftime('%m/%d/%Y')
                    temp_list.append(date_format)
                else:
                    temp_list.append(row.value)
        if len(temp_list) != 0:
            list_of_rows.append(temp_list)
    print(list_of_rows)
    return list_of_rows
