import openpyxl.utils.exceptions as openpyxl_exceptions
from selenium import webdriver
from openpyxl import load_workbook


# driver = webdriver.Chrome()


def entrypoint():
    wb = read_file()
    while wb is None:
        print("read_file failed, try again...")
        wb = read_file()
    print_sheets(wb)
    sheet = select_sheet(wb)
    while sheet is None:
        sheet = select_sheet(wb)
    parse_for_data(sheet)


def read_file():
    """
    Prints a prompt for user
    Receives string input
    Validates input
    Returns
    """
    print("Enter the name of the mileage sheet. Don't forget the extension: ")
    user_input_filename = input()
    try:
        workbook = load_workbook(filename=user_input_filename)
    except FileNotFoundError:
        print("File not found")
        return None
    except openpyxl_exceptions.InvalidFileException:
        print("Unsupported file format")
        return None
    else:
        print("File loaded successfully")
        return workbook


def print_sheets(workbook):
    print('Mileage sheets: ')
    print(workbook.sheetnames)


def select_sheet(workbook):
    print("Which sheet would you like to use?")
    user_input = input()
    if user_input in workbook.sheetnames:
        print("Copying " + user_input)
        current_worksheet = workbook[user_input]
        return current_worksheet
    else:
        print("The sheet: " + user_input + " was not found.")
        return None


def parse_for_data(current_worksheet):
    current_cell = current_worksheet['B8']
    print(current_cell.value)
    pass


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    entrypoint()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
